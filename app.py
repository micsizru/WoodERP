import os
import io
import calendar
from datetime import datetime, date, timedelta

from dateutil.relativedelta import relativedelta
from flask import (
    Flask, render_template, request, redirect,
    url_for, flash, jsonify, send_file
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, text
import pandas as pd
import pytz
import pdfkit
import platform

# ───────────────────────── App Konfigürasyonu ─────────────────────────
app = Flask(__name__)
app.config["SECRET_KEY"] = "marmara-aydoganlar-secret-key-2026"
basedir = os.path.abspath(os.path.dirname(__file__))
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(basedir, "veritabani.db")

# ───────────────────────── PDF Konfigürasyonu (pdfkit) ─────────────────────────
try:
    if platform.system() == "Windows":
        # Windows için tipik yol (Kullanıcı kendi kurulumuna göre düzenlemeli)
        path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
        pdf_config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf) if os.path.exists(path_wkhtmltopdf) else None
    else:
        # PythonAnywhere / Linux için tipik yol
        path_wkhtmltopdf = '/usr/bin/wkhtmltopdf'
        pdf_config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf) if os.path.exists(path_wkhtmltopdf) else None
except Exception:
    pdf_config = None
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# ───────────────────────── Veritabanı Modelleri ─────────────────────────

class Fis(db.Model):
    """Sevkiyat üst bilgisi (kamyon/fiş)."""
    __tablename__ = "fis"

    id = db.Column(db.Integer, primary_key=True)
    tarih = db.Column(db.Date, nullable=False, default=lambda: get_istanbul_time().date())
    sevk_eden_cari = db.Column(db.String(200), nullable=False)
    sevk_yeri_fabrika = db.Column(db.String(200), nullable=False)
    sevk_yeri_fis_no = db.Column(db.String(100), nullable=False)
    plaka_no = db.Column(db.String(50), nullable=False)

    # İlişki: Bir fişin birden fazla detay satırı olabilir
    detaylar = db.relationship(
        "FisDetayi", backref="fis", lazy=True, cascade="all, delete-orphan"
    )

    def __repr__(self):
        return f"<Fis {self.id} – {self.plaka_no}>"


class FisDetayi(db.Model):
    """Kamyonun içeriği (her bir kalem satırı)."""
    __tablename__ = "fis_detayi"

    id = db.Column(db.Integer, primary_key=True)
    fis_id = db.Column(db.Integer, db.ForeignKey("fis.id"), nullable=False)
    agac_cinsi = db.Column(db.String(100), nullable=False)
    cap = db.Column(db.String(50), nullable=False)
    miktar = db.Column(db.Float, nullable=False)
    birim = db.Column(db.String(20), nullable=False)
    birim_fiyat = db.Column(db.Float, nullable=False)
    toplam_tutar = db.Column(db.Float, nullable=False)

    def __repr__(self):
        return f"<FisDetayi {self.id} – {self.agac_cinsi}>"


class Cari(db.Model):
    """Cari (Müşteri/Tedarikçi) kartları."""
    __tablename__ = "cari"

    firma_kodu = db.Column(db.String(20), primary_key=True)
    firma_adi = db.Column(db.String(200), unique=True, nullable=False)
    vergi_dairesi = db.Column(db.String(100))
    vergi_numarasi = db.Column(db.String(50))
    firma_adres = db.Column(db.Text)
    e_posta = db.Column(db.String(100))
    telefon_no = db.Column(db.String(20))

    def __repr__(self):
        return f"<Cari {self.firma_kodu} - {self.firma_adi}>"


class Fabrika(db.Model):
    """Fabrika (Sevk Yeri) kartları."""
    __tablename__ = "fabrika"

    firma_kodu = db.Column(db.String(20), primary_key=True)
    firma_adi = db.Column(db.String(200), unique=True, nullable=False)
    vergi_dairesi = db.Column(db.String(100))
    vergi_numarasi = db.Column(db.String(50))
    firma_adres = db.Column(db.Text)
    e_posta = db.Column(db.String(100))
    telefon_no = db.Column(db.String(20))

    def __repr__(self):
        return f"<Fabrika {self.firma_kodu} - {self.firma_adi}>"


# ───────────────────────── Sabit Dropdown Verileri ─────────────────────────

AGAC_CINSLERI = [
    "Kayın Odun",
    "Çam Odun",
    "Çam Odun (Yanık Saha)",
    "Meşe Odun",
    "Kavak Odun",
    "Orman Kavağı",
    "Kestane Odun",
    "Çınar Odun",
    "Söğüt.Kib.Göb.Selvi Kavak",
    "Kızılağaç-Akasya",
    "Ihlamur",
    "Diş Budak",
]

#Burada daha fazla ağaç cinsi eklenebilir sonradan


CAPLER = ["-", "6cm altı", "6-15cm arası", "15 üstü"] #Çaplar'ı capler olarak yazmış, tüm değişkenlerin adlarını değiştirmeye üşendim

BIRIMLER = ["Ton", "Ster", "m³", "Kg"]

TURKCE_AYLAR = [
    "", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]


def tarih_araligi_hesapla(filtre, baslangic_str=None, bitis_str=None):
    """Filtre parametresine göre (başlangıç, bitiş) tarih çifti döndürür."""
    bugun = date.today()

    if filtre == "bugun":
        return bugun, bugun

    elif filtre == "son_3_gun":
        return bugun - timedelta(days=2), bugun

    elif filtre == "son_hafta":
        return bugun - timedelta(days=6), bugun

    elif filtre == "bu_hafta":
        pazartesi = bugun - timedelta(days=bugun.weekday())
        return pazartesi, bugun

    elif filtre == "onceki_hafta":
        bu_pazartesi = bugun - timedelta(days=bugun.weekday())
        onceki_pazartesi = bu_pazartesi - timedelta(days=7)
        onceki_pazar = bu_pazartesi - timedelta(days=1)
        return onceki_pazartesi, onceki_pazar

    elif filtre == "son_ay":
        bir_ay_once = bugun - relativedelta(months=1)
        return bir_ay_once, bugun

    elif filtre and filtre.startswith("yil_"):
        yil = int(filtre.split("_")[1])
        return date(yil, 1, 1), date(yil, 12, 31)

    elif filtre and filtre.startswith("ay_"):
        parcalar = filtre.split("_")  # ay_2026_03
        yil, ay = int(parcalar[1]), int(parcalar[2])
        son_gun = calendar.monthrange(yil, ay)[1]
        return date(yil, ay, 1), date(yil, ay, son_gun)

    elif filtre and filtre.startswith("hafta_"):
        parcalar = filtre.split("_")  # hafta_2026_03_1
        yil, ay, hafta_no = int(parcalar[1]), int(parcalar[2]), int(parcalar[3])
        # Ayın ilk gününün pazartesisini bul
        ayin_ilk_gunu = date(yil, ay, 1)
        ilk_pazartesi_offset = (7 - ayin_ilk_gunu.weekday()) % 7
        if ayin_ilk_gunu.weekday() == 0:  # Zaten pazartesi
            ilk_pazartesi = ayin_ilk_gunu
        else:
            ilk_pazartesi = ayin_ilk_gunu + timedelta(days=ilk_pazartesi_offset)
        hafta_baslangic = ilk_pazartesi + timedelta(weeks=hafta_no - 1)
        hafta_bitis = hafta_baslangic + timedelta(days=6)
        # Ayın son gününü aşmasın
        son_gun = calendar.monthrange(yil, ay)[1]
        ayin_sonu = date(yil, ay, son_gun)
        hafta_bitis = min(hafta_bitis, ayin_sonu)
        return hafta_baslangic, hafta_bitis

    elif filtre == "ozel" and baslangic_str and bitis_str:
        baslangic = datetime.strptime(baslangic_str, "%Y-%m-%d").date()
        bitis = datetime.strptime(bitis_str, "%Y-%m-%d").date()
        return baslangic, bitis

    # Filtre yok veya geçersiz → tüm veriler
    return None, None


def get_istanbul_time():
    """Türkiye yerel saatini döner."""
    return datetime.now(pytz.timezone('Europe/Istanbul'))

def istanbul_simdi():
    """Türkiye yerel saatini döner."""
    return get_istanbul_time()


def yeni_firma_kodu_uret(tur):
    """Cari için C-001, Fabrika için F-001 formatında otomatik kod üretir."""
    if tur == 'cari':
        son_kayit = Cari.query.order_by(Cari.firma_kodu.desc()).first()
        prefix = 'C-'
    else:
        son_kayit = Fabrika.query.order_by(Fabrika.firma_kodu.desc()).first()
        prefix = 'F-'
        
    if son_kayit and son_kayit.firma_kodu.startswith(prefix):
        try:
            num = int(son_kayit.firma_kodu.split('-')[1])
            yeni_num = num + 1
        except ValueError:
            yeni_num = 1
    else:
        yeni_num = 1
        
    return f"{prefix}{yeni_num:03d}"


# ───────────────────────── Route'lar ─────────────────────────

@app.route("/")
def index():
    """Dashboard – toplam sayılar ve son 5 fiş."""
    toplam_fis = Fis.query.count()
    toplam_kalem = FisDetayi.query.count()
    son_fisler = Fis.query.order_by(Fis.tarih.desc(), Fis.id.desc()).limit(5).all()
    return render_template("index.html", toplam_fis=toplam_fis,
                           toplam_kalem=toplam_kalem, son_fisler=son_fisler)


@app.route("/yeni_fis", methods=["GET"])
def yeni_fis_formu():
    """Yeni fiş oluşturma formu."""
    bugun = get_istanbul_time().strftime("%Y-%m-%d")
    cariler = Cari.query.order_by(Cari.firma_adi).all()
    fabrikalar = Fabrika.query.order_by(Fabrika.firma_adi).all()
    
    return render_template(
        "yeni_fis.html",
        bugun=bugun,
        agac_cinsleri=AGAC_CINSLERI,
        capler=CAPLER,
        birimler=BIRIMLER,
        cariler=cariler,
        fabrikalar=fabrikalar
    )


@app.route("/yeni_fis", methods=["POST"])
def yeni_fis_kaydet():
    """JSON verisini alıp Fis ve FisDetayi tablolarına kaydeder."""
    try:
        veri = request.get_json()

        # Üst bilgi
        tarih_str = veri.get("tarih", date.today().isoformat())
        tarih = datetime.strptime(tarih_str, "%Y-%m-%d").date()

        yeni_fis = Fis(
            tarih=tarih,
            sevk_eden_cari=veri["sevk_eden_cari"].strip(),
            sevk_yeri_fabrika=veri["sevk_yeri_fabrika"].strip(),
            sevk_yeri_fis_no=veri["sevk_yeri_fis_no"].strip(),
            plaka_no=veri["plaka_no"].strip(),
        )
        db.session.add(yeni_fis)
        db.session.flush()  # id ataması için

        # Detay satırları
        kalemler = veri.get("kalemler", [])
        if not kalemler:
            return jsonify({"durum": "hata", "mesaj": "En az bir kalem satırı eklemelisiniz."}), 400

        for k in kalemler:
            miktar = float(k["miktar"])
            birim_fiyat = float(k["birim_fiyat"])
            toplam_tutar = round(miktar * birim_fiyat, 2)

            detay = FisDetayi(
                fis_id=yeni_fis.id,
                agac_cinsi=k["agac_cinsi"],
                cap=k.get("cap", "-") or "-",
                miktar=miktar,
                birim=k["birim"],
                birim_fiyat=birim_fiyat,
                toplam_tutar=toplam_tutar,
            )
            db.session.add(detay)

        db.session.commit()
        return jsonify({"durum": "basarili", "mesaj": "Fiş başarıyla kaydedildi.", "fis_id": yeni_fis.id})

    except Exception as e:
        db.session.rollback()
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


@app.route("/fisleri_goruntule")
def fisleri_goruntule():
    """Kayıtlı tüm fişleri tarih sırasıyla listeler."""
    fisler = Fis.query.order_by(Fis.tarih.desc(), Fis.id.desc()).all()
    return render_template("fisleri_goruntule.html", fisler=fisler)


@app.route("/fis_detay/<int:fis_id>")
def fis_detay(fis_id):
    """Tek bir fişin detay sayfası."""
    fis = Fis.query.get_or_404(fis_id)
    genel_toplam = sum(d.toplam_tutar for d in fis.detaylar)
    return render_template("fis_detay.html", fis=fis, genel_toplam=genel_toplam)


@app.route("/fis_sil/<int:fis_id>")
def fis_sil(fis_id):
    """Fişi ve tüm detay satırlarını siler."""
    fis = Fis.query.get_or_404(fis_id)
    db.session.delete(fis)
    db.session.commit()
    flash("Fiş başarıyla silindi.", "success")
    return redirect(url_for("fisleri_goruntule"))


@app.route("/tek_fis_excel/<int:fis_id>")
def tek_fis_excel(fis_id):
    """Belirli bir fişi şık formatlı Excel olarak indirir."""
    fis = Fis.query.get_or_404(fis_id)
    
    # Verileri DataFrame için hazırla
    veriler = []
    for d in fis.detaylar:
        veriler.append({
            "Fiş No": fis.id,
            "Tarih": fis.tarih,
            "Sevk Eden Cari": fis.sevk_eden_cari,
            "Fabrika": fis.sevk_yeri_fabrika,
            "Sevkiyat Fiş No": fis.sevk_yeri_fis_no,
            "Plaka No": fis.plaka_no,
            "Ağaç Cinsi": d.agac_cinsi,
            "Çap": d.cap,
            "Miktar": d.miktar,
            "Birim": d.birim,
            "Birim Fiyat": d.birim_fiyat,
            "Toplam Tutar": d.toplam_tutar,
        })
    
    df = pd.DataFrame(veriler)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=f"Fis_{fis.id}")
        ws = writer.sheets[f"Fis_{fis.id}"]

        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Başlık satırı stili
        header_fill = PatternFill("solid", fgColor="1B5E20")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Zebra striping
        stripe_fill = PatternFill("solid", fgColor="F5F5F0")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = stripe_fill

        # Auto-fit sütunlar
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
                except Exception: pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    output.seek(0)
    dosya_adi = f"Fis_{fis.id}_{fis.sevk_eden_cari.replace(' ', '_')}_{fis.tarih}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=dosya_adi,
    )


@app.route("/tek_fis_pdf/<int:fis_id>")
def tek_fis_pdf(fis_id):
    """Belirli bir fişi şık formatlı PDF olarak indirir."""
    fis = Fis.query.get_or_404(fis_id)
    
    # PDF için HTML içeriği oluştur
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'DejaVu Sans', sans-serif; font-size: 12px; color: #333; }}
            .header {{ text-align: center; margin-bottom: 20px; }}
            .info-table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            .info-table td {{ padding: 8px; border: 1px solid #ddd; }}
            .info-label {{ font-weight: bold; background-color: #f9f9f9; width: 30%; }}
            .items-table {{ width: 100%; border-collapse: collapse; }}
            .items-table th {{ background-color: #1B5E20; color: white; padding: 10px; text-align: center; }}
            .items-table td {{ padding: 8px; border: 1px solid #ddd; text-align: center; }}
            .zebra-row {{ background-color: #F5F5F0; }}
            .footer {{ margin-top: 30px; text-align: right; font-style: italic; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1 style="color: #1B5E20; margin-bottom: 5px;">Marmara Aydoğanlar</h1>
            <h3 style="margin-top: 5px;">Sevkiyat Fişi Bilgileri</h3>
        </div>

        <table class="info-table">
            <tr>
                <td class="info-label">Fiş No:</td><td>#{fis.id}</td>
                <td class="info-label">Tarih:</td><td>{fis.tarih.strftime('%d.%m.%Y')}</td>
            </tr>
            <tr>
                <td class="info-label">Sevk Eden Cari:</td><td>{fis.sevk_eden_cari}</td>
                <td class="info-label">Plaka No:</td><td>{fis.plaka_no}</td>
            </tr>
            <tr>
                <td class="info-label">S. Yeri / Fabrika:</td><td>{fis.sevk_yeri_fabrika}</td>
                <td class="info-label">S. Yeri Fiş No:</td><td>{fis.sevk_yeri_fis_no}</td>
            </tr>
        </table>

        <table class="items-table">
            <thead>
                <tr>
                    <th>Ağaç Cinsi</th>
                    <th>Çap</th>
                    <th>Miktar</th>
                    <th>Birim</th>
                    <th>Birim Fiyat</th>
                    <th>Toplam Tutar</th>
                </tr>
            </thead>
            <tbody>
    """
    
    toplam_genel = 0
    for i, d in enumerate(fis.detaylar):
        row_class = 'class="zebra-row"' if i % 2 == 1 else ''
        toplam_genel += d.toplam_tutar
        html_content += f"""
            <tr {row_class}>
                <td>{d.agac_cinsi}</td>
                <td>{d.cap}</td>
                <td>{d.miktar}</td>
                <td>{d.birim}</td>
                <td>{d.birim_fiyat} ₺</td>
                <td>{d.toplam_tutar} ₺</td>
            </tr>
        """
        
    html_content += f"""
            </tbody>
            <tfoot>
                <tr style="font-weight: bold; background-color: #eee;">
                    <td colspan="5" style="text-align: right; padding: 10px;">Genel Toplam:</td>
                    <td style="text-align: center; padding: 10px;">{toplam_genel:.2f} ₺</td>
                </tr>
            </tfoot>
        </table>

        <div class="footer">
            <p>Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
        </div>
    </body>
    </html>
    """

    options = {
        'encoding': 'UTF-8',
        'quiet': '',
        'no-outline': None,
        'margin-top': '0.75in',
        'margin-right': '0.75in',
        'margin-bottom': '0.75in',
        'margin-left': '0.75in',
    }

    try:
        pdf_bytes = pdfkit.from_string(html_content, False, options=options, configuration=pdf_config)
        
        output = io.BytesIO(pdf_bytes)
        dosya_adi = f"Fis_{fis.id}_{fis.sevk_eden_cari.replace(' ', '_')}.pdf"
        
        return send_file(
            output,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=dosya_adi
        )
    except Exception as e:
        flash(f"PDF oluşturulurken hata oluştu: {str(e)}", "danger")
        return redirect(url_for("fis_detay", fis_id=fis.id))


@app.route("/fis_duzenle/<int:fis_id>", methods=["GET", "POST"])
def fis_duzenle(fis_id):
    """Mevcut bir fişi düzenler."""
    fis = Fis.query.get_or_404(fis_id)
    
    if request.method == "POST":
        try:
            veri = request.get_json()
            
            # Üst bilgileri güncelle
            tarih_str = veri.get("tarih")
            if tarih_str:
                fis.tarih = datetime.strptime(tarih_str, "%Y-%m-%d").date()
            
            fis.sevk_eden_cari = veri["sevk_eden_cari"].strip()
            fis.sevk_yeri_fabrika = veri["sevk_yeri_fabrika"].strip()
            fis.sevk_yeri_fis_no = veri["sevk_yeri_fis_no"].strip()
            fis.plaka_no = veri["plaka_no"].strip()
            
            # Kalemleri güncelle (Eskileri sil, yenileri ekle - En güvenli yol)
            FisDetayi.query.filter_by(fis_id=fis.id).delete()
            
            kalemler = veri.get("kalemler", [])
            for k in kalemler:
                miktar = float(k["miktar"])
                birim_fiyat = float(k["birim_fiyat"])
                toplam_tutar = round(miktar * birim_fiyat, 2)
                
                yeni_detay = FisDetayi(
                    fis_id=fis.id,
                    agac_cinsi=k["agac_cinsi"],
                    cap=k.get("cap", "-") or "-",
                    miktar=miktar,
                    birim=k["birim"],
                    birim_fiyat=birim_fiyat,
                    toplam_tutar=toplam_tutar
                )
                db.session.add(yeni_detay)
                
            db.session.commit()
            return jsonify({"durum": "basarili", "mesaj": "Fiş başarıyla güncellendi."})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({"durum": "hata", "mesaj": str(e)}), 500

    # GET: Sayfayı render et
    cariler = Cari.query.order_by(Cari.firma_adi).all()
    fabrikalar = Fabrika.query.order_by(Fabrika.firma_adi).all()

    # SQLAlchemy objelerini JSON serializable listeye çevir
    detaylar_list = []
    for d in fis.detaylar:
        detaylar_list.append({
            "agac_cinsi": d.agac_cinsi,
            "cap": d.cap,
            "miktar": d.miktar,
            "birim": d.birim,
            "birim_fiyat": d.birim_fiyat,
            "toplam_tutar": d.toplam_tutar
        })
    
    return render_template(
        "fis_duzenle.html",
        fis=fis,
        detaylar_list=detaylar_list,
        agac_cinsleri=AGAC_CINSLERI,
        capler=CAPLER,
        birimler=BIRIMLER,
        cariler=cariler,
        fabrikalar=fabrikalar
    )


# ─── Kartlar (Cari & Fabrika) Route'ları ───

@app.route("/kartlar")
def kartlar():
    """Cari ve Fabrika listeleme sayfası + Ticaret Hacimleri."""
    now = istanbul_simdi()
    bu_ay_baslangic = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    son_30_gun_baslangic = (now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)

    def hacim_hesapla(firma_adi, tip='cari'):
        # Fis ve FisDetayi birleştirerek toplam tutar hesabı
        sorgu_base = db.session.query(db.func.sum(FisDetayi.toplam_tutar)).join(Fis)
        if tip == 'cari':
            sorgu_base = sorgu_base.filter(Fis.sevk_eden_cari == firma_adi)
        else:
            sorgu_base = sorgu_base.filter(Fis.sevk_yeri_fabrika == firma_adi)
        
        # 1. Tüm Zamanlar
        tum_zamanlar = sorgu_base.scalar() or 0.0
        
        # 2. Bu Ay
        bu_ay = sorgu_base.filter(Fis.tarih >= bu_ay_baslangic.date()).scalar() or 0.0
        
        # 3. Son 30 Gün
        son_30 = sorgu_base.filter(Fis.tarih >= son_30_gun_baslangic.date()).scalar() or 0.0
        
        return {
            'tum': round(tum_zamanlar, 2),
            'bu_ay': round(bu_ay, 2),
            'son_30': round(son_30, 2)
        }

    cariler = Cari.query.order_by(Cari.firma_kodu).all()
    cariler_data = []
    for c in cariler:
        c.hacim = hacim_hesapla(c.firma_adi, 'cari')
        cariler_data.append(c)

    fabrikalar = Fabrika.query.order_by(Fabrika.firma_kodu).all()
    fabrikalar_data = []
    for f in fabrikalar:
        f.hacim = hacim_hesapla(f.firma_adi, 'fabrika')
        fabrikalar_data.append(f)

    return render_template("kartlar.html", cariler=cariler_data, fabrikalar=fabrikalar_data)


@app.route("/kart_ekle/<tur>", methods=["POST"])
def kart_ekle(tur):
    """Yeni Cari veya Fabrika ekler."""
    Model = Cari if tur == 'cari' else Fabrika
    
    firma_adi = request.form.get("firma_adi", "").strip()
    if not firma_adi:
        flash("Firma adı zorunludur.", "danger")
        return redirect(url_for("kartlar"))
        
    mevcut = Model.query.filter_by(firma_adi=firma_adi).first()
    if mevcut:
        flash(f"Bu firma adı zaten mevcut: {firma_adi}", "danger")
        return redirect(url_for("kartlar"))
        
    yeni_kod = yeni_firma_kodu_uret(tur)
    
    yeni_kart = Model(
        firma_kodu=yeni_kod,
        firma_adi=firma_adi,
        vergi_dairesi=request.form.get("vergi_dairesi", "").strip(),
        vergi_numarasi=request.form.get("vergi_numarasi", "").strip(),
        firma_adres=request.form.get("firma_adres", "").strip(),
        e_posta=request.form.get("e_posta", "").strip(),
        telefon_no=request.form.get("telefon_no", "").strip()
    )
    db.session.add(yeni_kart)
    db.session.commit()
    flash(f"{'Cari' if tur == 'cari' else 'Fabrika'} başarıyla eklendi. (Kod: {yeni_kod})", "success")
    return redirect(url_for("kartlar"))


@app.route("/kart_duzenle/<tur>/<firma_kodu>", methods=["POST"])
def kart_duzenle(tur, firma_kodu):
    """Mevcut Cari veya Fabrikayı günceller."""
    Model = Cari if tur == 'cari' else Fabrika
    kart = Model.query.get_or_404(firma_kodu)
    
    yeni_ad = request.form.get("firma_adi", "").strip()
    if not yeni_ad:
        flash("Firma adı zorunludur.", "danger")
        return redirect(url_for("kartlar"))
        
    mevcut = Model.query.filter(Model.firma_adi == yeni_ad, Model.firma_kodu != firma_kodu).first()
    if mevcut:
        flash(f"Bu firma adı başka bir kayıtta kullanılıyor: {yeni_ad}", "danger")
        return redirect(url_for("kartlar"))
        
    kart.firma_adi = yeni_ad
    kart.vergi_dairesi = request.form.get("vergi_dairesi", "").strip()
    kart.vergi_numarasi = request.form.get("vergi_numarasi", "").strip()
    kart.firma_adres = request.form.get("firma_adres", "").strip()
    kart.e_posta = request.form.get("e_posta", "").strip()
    kart.telefon_no = request.form.get("telefon_no", "").strip()
    
    db.session.commit()
    flash(f"{'Cari' if tur == 'cari' else 'Fabrika'} başarıyla güncellendi.", "success")
    return redirect(url_for("kartlar"))


@app.route("/kart_sil/<tur>/<firma_kodu>", methods=["POST"])
def kart_sil(tur, firma_kodu):
    """Mevcut Cari veya Fabrikayı siler."""
    Model = Cari if tur == 'cari' else Fabrika
    kart = Model.query.get_or_404(firma_kodu)
    db.session.delete(kart)
    db.session.commit()
    flash(f"{'Cari' if tur == 'cari' else 'Fabrika'} başarıyla silindi.", "success")
    return redirect(url_for("kartlar"))


@app.route("/raporlar")
def raporlar():
    """Rapor filtreleme sayfası."""
    bugun = date.today()
    # Veritabanındaki en eski yılı bul
    en_eski = db.session.query(db.func.min(Fis.tarih)).scalar()
    en_eski_yil = en_eski.year if en_eski else bugun.year
    yillar = list(range(bugun.year, en_eski_yil - 1, -1))

    # Filtreler için Cari ve Fabrika listesi
    cariler = Cari.query.order_by(Cari.firma_adi).all()
    fabrikalar = Fabrika.query.order_by(Fabrika.firma_adi).all()

    return render_template(
        "raporlar.html",
        bugun=bugun.isoformat(),
        yillar=yillar,
        aylar=TURKCE_AYLAR,
        cariler=cariler,
        fabrikalar=fabrikalar
    )


@app.route("/rapor_indir")
def rapor_indir():
    """Pandas ile JOIN yapıp filtrelenmiş Excel dosyası oluşturur ve indirir."""
    filtre = request.args.get("filtre", "")
    baslangic_str = request.args.get("baslangic", "")
    bitis_str = request.args.get("bitis", "")
    
    # Yeni filtreler
    cari_adi = request.args.get("cari_adi", "")
    fabrika_adi = request.args.get("fabrika_adi", "")

    baslangic, bitis = tarih_araligi_hesapla(filtre, baslangic_str, bitis_str)

    sorgu = (
        db.session.query(
            Fis.id.label("Fiş No"),
            Fis.tarih.label("Tarih"),
            Fis.sevk_eden_cari.label("Sevk Eden Cari"),
            Fis.sevk_yeri_fabrika.label("Fabrika"),
            Fis.sevk_yeri_fis_no.label("Sevkiyat Fiş No"),
            Fis.plaka_no.label("Plaka No"),
            FisDetayi.agac_cinsi.label("Ağaç Cinsi"),
            FisDetayi.cap.label("Çap"),
            FisDetayi.miktar.label("Miktar"),
            FisDetayi.birim.label("Birim"),
            FisDetayi.birim_fiyat.label("Birim Fiyat"),
            FisDetayi.toplam_tutar.label("Toplam Tutar"),
        )
        .join(FisDetayi, Fis.id == FisDetayi.fis_id)
    )

    # Tarih filtresini uygula
    if baslangic and bitis:
        sorgu = sorgu.filter(Fis.tarih >= baslangic, Fis.tarih <= bitis)

    # FİRMA FİLTRELERİ (God Mode değilse her durumda uygulanmalı)
    if filtre != 'god_mode':
        if cari_adi:
            sorgu = sorgu.filter(Fis.sevk_eden_cari == cari_adi)
        if fabrika_adi:
            sorgu = sorgu.filter(Fis.sevk_yeri_fabrika == fabrika_adi)

    sonuclar = sorgu.order_by(Fis.tarih.desc(), Fis.id.desc()).all()

    df = pd.DataFrame(sonuclar, columns=[
        "Fiş No", "Tarih", "Sevk Eden Cari", "Fabrika",
        "Sevkiyat Fiş No", "Plaka No", "Ağaç Cinsi", "Çap",
        "Miktar", "Birim", "Birim Fiyat", "Toplam Tutar",
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sevkiyat Raporu")
        ws = writer.sheets["Sevkiyat Raporu"]

        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Başlık satırı: koyu yeşil zemin, beyaz+kalın yazı
        header_fill = PatternFill("solid", fgColor="1B5E20")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Zebra striping: çok açık gri - beyaz
        stripe_fill = PatternFill("solid", fgColor="F5F5F0")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = stripe_fill

        # Auto-fit sütun genişlikleri
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
                except Exception:
                    pass
            adjusted = min(max_len + 4, 50)  # en fazla 50 karakter genislik
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    output.seek(0)

    # Dosya adını tarih aralığına göre oluştur
    if baslangic and bitis:
        dosya_adi = f"Marmara_Aydoganlar_Rapor_{baslangic}_{bitis}.xlsx"
    else:
        tarih_str = datetime.now().strftime("%Y-%m-%d")
        dosya_adi = f"Marmara_Aydoganlar_Rapor_Tumu_{tarih_str}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=dosya_adi,
    )


@app.route("/rapor_indir_pdf")
def rapor_indir_pdf():
    """Genel raporu şık formatlı PDF olarak indirir."""
    filtre = request.args.get("filtre", "")
    baslangic_str = request.args.get("baslangic", "")
    bitis_str = request.args.get("bitis", "")
    
    cari_adi = request.args.get("cari_adi", "")
    fabrika_adi = request.args.get("fabrika_adi", "")

    baslangic, bitis = tarih_araligi_hesapla(filtre, baslangic_str, bitis_str)

    sorgu = (
        db.session.query(
            Fis.id,
            Fis.tarih,
            Fis.sevk_eden_cari,
            Fis.sevk_yeri_fabrika,
            Fis.sevk_yeri_fis_no,
            Fis.plaka_no,
            FisDetayi.agac_cinsi,
            FisDetayi.cap,
            FisDetayi.miktar,
            FisDetayi.birim,
            FisDetayi.birim_fiyat,
            FisDetayi.toplam_tutar,
        )
        .join(FisDetayi, Fis.id == FisDetayi.fis_id)
    )

    if baslangic and bitis:
        sorgu = sorgu.filter(Fis.tarih >= baslangic, Fis.tarih <= bitis)

    # FİRMA FİLTRELERİ (God Mode değilse query'e ekle)
    if filtre != 'god_mode':
        if cari_adi:
            sorgu = sorgu.filter(Fis.sevk_eden_cari == cari_adi)
        if fabrika_adi:
            sorgu = sorgu.filter(Fis.sevk_yeri_fabrika == fabrika_adi)

    sonuclar = sorgu.order_by(Fis.tarih.desc(), Fis.id.desc()).all()

    # PDF için HTML içeriği oluştur
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'DejaVu Sans', sans-serif; font-size: 10px; color: #333; }}
            .header {{ text-align: center; margin-bottom: 20px; }}
            .items-table {{ width: 100%; border-collapse: collapse; }}
            .items-table th {{ background-color: #1B5E20; color: white; padding: 6px; text-align: center; border: 1px solid #1B5E20; }}
            .items-table td {{ padding: 5px; border: 1px solid #ddd; text-align: center; }}
            .zebra-row {{ background-color: #F5F5F0; }}
            .footer {{ margin-top: 20px; text-align: right; font-style: italic; font-size: 9px; }}
            .filter-info {{ margin-bottom: 15px; font-weight: bold; color: #555; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1 style="color: #1B5E20; margin-bottom: 5px;">Marmara Aydoğanlar</h1>
            <h3 style="margin-top: 5px;">Genel Sevkiyat Raporu</h3>
        </div>

        <div class="filter-info">
            Kapsam: {baslangic if baslangic else 'Tümü'} - {bitis if bitis else 'Tümü'} <br>
            Cari: {cari_adi if cari_adi else 'Tümü'} | Fabrika: {fabrika_adi if fabrika_adi else 'Tümü'}
        </div>

        <table class="items-table">
            <thead>
                <tr>
                    <th>Fiş No</th>
                    <th>Tarih</th>
                    <th>Cari</th>
                    <th>Fabrika</th>
                    <th>Plaka</th>
                    <th>Cins</th>
                    <th>Çap</th>
                    <th>Miktar</th>
                    <th>Birim</th>
                    <th>Fiyat</th>
                    <th>Toplam</th>
                </tr>
            </thead>
            <tbody>
    """
    
    toplam_genel = 0
    for i, res in enumerate(sonuclar):
        row_class = 'class="zebra-row"' if i % 2 == 1 else ''
        toplam_genel += res.toplam_tutar
        tarih_format = res.tarih.strftime('%d.%m.%Y')
        html_content += f"""
            <tr {row_class}>
                <td>{res.id}</td>
                <td>{tarih_format}</td>
                <td>{res.sevk_eden_cari}</td>
                <td>{res.sevk_yeri_fabrika}</td>
                <td>{res.plaka_no}</td>
                <td>{res.agac_cinsi}</td>
                <td>{res.cap}</td>
                <td>{res.miktar}</td>
                <td>{res.birim}</td>
                <td>{res.birim_fiyat}</td>
                <td style="font-weight: bold;">{res.toplam_tutar:.2f}</td>
            </tr>
        """
        
    html_content += f"""
            </tbody>
            <tfoot>
                <tr style="font-weight: bold; background-color: #eee;">
                    <td colspan="10" style="text-align: right; padding: 10px;">Genel Rapor Toplamı:</td>
                    <td style="text-align: center; padding: 10px;">{toplam_genel:.2f} ₺</td>
                </tr>
            </tfoot>
        </table>

        <div class="footer">
            <p>Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
        </div>
    </body>
    </html>
    """

    options = {
        'encoding': 'UTF-8',
        'quiet': '',
        'orientation': 'Landscape',
        'margin-top': '0.5in',
        'margin-right': '0.5in',
        'margin-bottom': '0.5in',
        'margin-left': '0.5in',
    }

    try:
        pdf_bytes = pdfkit.from_string(html_content, False, options=options, configuration=pdf_config)
        
        output = io.BytesIO(pdf_bytes)
        dosya_adi = f"Marmara_Aydoganlar_Rapor_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        
        return send_file(
            output,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=dosya_adi
        )
    except Exception as e:
        flash(f"PDF oluşturulurken hata oluştu: {{str(e)}}", "danger")
        return redirect(url_for("raporlar"))


def check_and_update_tables():
    """Modellerdeki yeni sütunları veritabanına otomatik ekler (SQLite uyumlu)."""
    with app.app_context():
        inspector = inspect(db.engine)
        for table_name, table in db.metadata.tables.items():
            try:
                existing_columns = [c["name"] for c in inspector.get_columns(table_name)]
            except Exception:
                continue

            for column in table.columns:
                if column.name not in existing_columns:
                    # SQLite'da NOT NULL sütun eklerken DEFAULT değer zorunludur.
                    default_clause = ""
                    if not column.nullable:
                        col_type = str(column.type).upper()
                        if any(t in col_type for t in ["INT", "FLOAT", "NUMERIC", "DECIMAL"]):
                            default_clause = " DEFAULT 0"
                        elif "BOOL" in col_type:
                            default_clause = " DEFAULT 0"
                        else:
                            default_clause = " DEFAULT ''"
                    
                    sql = f'ALTER TABLE {table_name} ADD COLUMN {column.name} {column.type}{default_clause}'
                    try:
                        db.session.execute(text(sql))
                        db.session.commit()
                        print(f"[*] Migrasyon: {table_name} tablosuna {column.name} sütunu eklendi.")
                    except Exception as e:
                        db.session.rollback()
                        print(f"[!] Migrasyon Hatası ({table_name}.{column.name}): {e}")

# ───────────────────────── Uygulama Başlatma ─────────────────────────

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        check_and_update_tables()
    app.run(debug=True, host="0.0.0.0", port=5000)





