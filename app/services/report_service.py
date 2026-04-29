import io
from datetime import datetime
import pandas as pd
import pdfkit
from flask import send_file, flash

class ReportService:
    @staticmethod
    def generate_excel_for_fis(fis):
        veriler = []
        for d in fis.detaylar:
            veriler.append({
                "Fiş No": fis.id,
                "Tarih": fis.tarih,
                "Sevk Eden Cari": fis.guncel_sevk_eden_cari,
                "Fabrika": fis.guncel_sevk_yeri_fabrika,
                "Sevkiyat Fiş No": fis.sevk_yeri_fis_no,
                "Plaka No": fis.plaka_no,
                "Ağaç Cinsi": d.agac_cinsi,
                "Çap": d.cap,
                "Miktar": d.miktar,
                "Birim": d.birim,
                "Birim Fiyat": d.birim_fiyat,
                "Toplam Tutar": d.toplam_tutar,
            })
        
        df = pd.DataFrame(veriler)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=f"Fis_{fis.id}")
            ws = writer.sheets[f"Fis_{fis.id}"]

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            header_fill = PatternFill("solid", fgColor="1B5E20")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            stripe_fill = PatternFill("solid", fgColor="F5F5F0")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = stripe_fill

            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in column_cells:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                        max_len = max(max_len, len(val))
                    except Exception: pass
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        output.seek(0)
        dosya_adi = f"Fis_{fis.id}_{fis.guncel_sevk_eden_cari.replace(' ', '_')}_{fis.tarih}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=dosya_adi,
        )

    @staticmethod
    def generate_pdf_for_fis(fis, pdf_config):
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: 'DejaVu Sans', sans-serif; font-size: 12px; color: #333; }}
                .header {{ text-align: center; margin-bottom: 20px; }}
                .info-table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                .info-table td {{ padding: 8px; border: 1px solid #ddd; }}
                .info-label {{ font-weight: bold; background-color: #f9f9f9; width: 30%; }}
                .items-table {{ width: 100%; border-collapse: collapse; }}
                .items-table th {{ background-color: #1B5E20; color: white; padding: 10px; text-align: center; }}
                .items-table td {{ padding: 8px; border: 1px solid #ddd; text-align: center; }}
                .zebra-row {{ background-color: #F5F5F0; }}
                .footer {{ margin-top: 30px; text-align: right; font-style: italic; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1 style="color: #1B5E20; margin-bottom: 5px;">Marmara Aydoğanlar</h1>
                <h3 style="margin-top: 5px;">Sevkiyat Fişi Bilgileri</h3>
            </div>

            <table class="info-table">
                <tr>
                    <td class="info-label">Fiş No:</td><td>#{fis.id}</td>
                    <td class="info-label">Tarih:</td><td>{fis.tarih.strftime('%d.%m.%Y')}</td>
                </tr>
                <tr>
                    <td class="info-label">Sevk Eden Cari:</td><td>{fis.guncel_sevk_eden_cari}</td>
                    <td class="info-label">Plaka No:</td><td>{fis.plaka_no}</td>
                </tr>
                <tr>
                    <td class="info-label">S. Yeri / Fabrika:</td><td>{fis.guncel_sevk_yeri_fabrika}</td>
                    <td class="info-label">S. Yeri Fiş No:</td><td>{fis.sevk_yeri_fis_no}</td>
                </tr>
            </table>

            <table class="items-table">
                <thead>
                    <tr>
                        <th>Ağaç Cinsi</th>
                        <th>Çap</th>
                        <th>Miktar</th>
                        <th>Birim</th>
                        <th>Birim Fiyat</th>
                        <th>Toplam Tutar</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        toplam_genel = 0
        for i, d in enumerate(fis.detaylar):
            row_class = 'class="zebra-row"' if i % 2 == 1 else ''
            toplam_genel += d.toplam_tutar
            html_content += f"""
                <tr {row_class}>
                    <td>{d.agac_cinsi}</td>
                    <td>{d.cap}</td>
                    <td>{d.miktar}</td>
                    <td>{d.birim}</td>
                    <td>{d.birim_fiyat} ₺</td>
                    <td>{d.toplam_tutar} ₺</td>
                </tr>
            """
            
        html_content += f"""
                </tbody>
                <tfoot>
                    <tr style="font-weight: bold; background-color: #eee;">
                        <td colspan="5" style="text-align: right; padding: 10px;">Genel Toplam:</td>
                        <td style="text-align: center; padding: 10px;">{toplam_genel:.2f} ₺</td>
                    </tr>
                </tfoot>
            </table>

            <div class="footer">
                <p>Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
            </div>
        </body>
        </html>
        """

        options = {
            'encoding': 'UTF-8',
            'quiet': '',
            'no-outline': None,
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
        }

        try:
            pdf_bytes = pdfkit.from_string(html_content, False, options=options, configuration=pdf_config)
            output = io.BytesIO(pdf_bytes)
            dosya_adi = f"Fis_{fis.id}_{fis.guncel_sevk_eden_cari.replace(' ', '_')}.pdf"
            
            return send_file(
                output,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=dosya_adi
            )
        except Exception as e:
            return None, str(e)
